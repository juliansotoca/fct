import os
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers


HEADER_FONT = Font(name="Arial", bold=True, size=14, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(name="Arial", bold=True, size=12, color="2F5496")
BOLD_FONT = Font(name="Arial", bold=True, size=10)
NORMAL_FONT = Font(name="Arial", size=10)
POSITIVE_FONT = Font(name="Arial", size=10, color="006100")
NEGATIVE_FONT = Font(name="Arial", size=10, color="9C0006")
GAINS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOSSES_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="2F5496")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def format_eur(value):
    if value is None:
        return ""
    return f"{value:,.2f} €"


def create_report(fifo_by_year, inventory=None, output_dir="reports"):
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(output_dir, f"tax_report_{timestamp}.xlsx")

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Resumen"
    _create_summary_sheet(summary_ws, fifo_by_year)

    for year, data in sorted(fifo_by_year.items()):
        ws = wb.create_sheet(title=str(year))
        _create_year_sheet(ws, year, data)

    if inventory:
        ws = wb.create_sheet(title="Inventario Actual")
        _create_inventory_sheet(ws, inventory)

    wb.save(filename)
    return filename


def _set_header_row(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _set_subheader_row(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_cell(cell, bold=False, number_format=False):
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center" if number_format else "left")


def _create_summary_sheet(ws, fifo_by_year):
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="INFORME FISCAL CRIPTOMONEDAS - ESPAÑA")
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    subtitle = ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Método: FIFO | Moneda: EUR")
    subtitle.font = NORMAL_FONT
    subtitle.alignment = Alignment(horizontal="center")

    row = 4
    _set_header_row(ws, row, ["Año", "Ganancias Patrimoniales", "Pérdidas Patrimoniales", "Ganancia/Pérdida Neta", "Rendimientos Capital Mobiliario", "Impuesto Estimado"])

    total_gains = Decimal("0")
    total_losses = Decimal("0")
    total_net = Decimal("0")
    total_income = Decimal("0")

    for year, data in sorted(fifo_by_year.items()):
        row += 1
        gains = data["total_gains"]
        losses = data["total_losses"]
        net = data["net_gain_loss"]
        income = data["income_value_eur"]
        gain_tax, income_tax = _compute_spanish_tax(net, income)
        total_tax = gain_tax + income_tax

        values = [year, gains, losses, net, income, total_tax]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col)
            if col == 1:
                cell.value = val
                cell.font = BOLD_FONT
            else:
                cell.value = float(val) if isinstance(val, Decimal) else val
                cell.number_format = '#,##0.00 "€"'
                if col == 4:
                    cell.font = POSITIVE_FONT if val >= 0 else NEGATIVE_FONT
                    cell.fill = GAINS_FILL if val >= 0 else LOSSES_FILL
                elif col == 6:
                    cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        total_gains += gains
        total_losses += losses
        total_net += net
        total_income += income

    row += 2
    ws.merge_cells(f"A{row}:B{row}")
    total_label = ws.cell(row=row, column=1, value="TOTALES ACUMULADOS")
    total_label.font = Font(name="Arial", bold=True, size=12, color="2F5496")

    row += 1
    total_gain_tax, total_income_tax = _compute_spanish_tax(total_net, total_income)
    totals = [
        ("Ganancias totales:", total_gains),
        ("Pérdidas totales:", total_losses),
        ("Ganancia/Pérdida neta:", total_net),
        ("Rendimientos capital mobiliario:", total_income),
        ("Impuesto estimado total:", total_gain_tax + total_income_tax),
    ]

    for label, value in totals:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=row, column=2, value=float(value))
        cell.number_format = '#,##0.00 "€"'
        cell.font = BOLD_FONT
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    note = ws.cell(row=row, column=1, value="NOTA: Este informe es orientativo. Consulta con un asesor fiscal para la declaración oficial.")
    note.font = Font(name="Arial", italic=True, size=9, color="808080")

    _add_tax_bracket_info(ws, row + 2)


def _add_tax_bracket_info(ws, start_row):
    ws.merge_cells(f"A{start_row}:F{start_row}")
    ws.cell(row=start_row, column=1, value="TRAMOS IRPF 2024 - RENTAS DEL AHORRO (España)").font = Font(name="Arial", bold=True, size=10, color="2F5496")

    brackets = [
        ("Hasta 6.000 €", "19%"),
        ("De 6.000 € a 50.000 €", "21%"),
        ("De 50.000 € a 200.000 €", "23%"),
        ("De 200.000 € a 2.000.000 €", "27%"),
        ("Más de 2.000.000 €", "28%"),
    ]

    row = start_row + 1
    _set_subheader_row(ws, row, ["Tramo", "Tipo impositivo"])
    for tramo, rate in brackets:
        row += 1
        ws.cell(row=row, column=1, value=tramo).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=rate).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")


def _create_year_sheet(ws, year, data):
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 40

    ws.merge_cells("A1:I1")
    title = ws.cell(row=1, column=1, value=f"EJERCICIO FISCAL {year}")
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center")

    row = 3
    _create_disposals_section(ws, row, data["disposals"])
    row = _get_next_row(ws) + 2

    _create_incomes_section(ws, row, data["incomes"])
    row = _get_next_row(ws) + 2

    _create_deposits_section(ws, row, data["deposits"])
    row = _get_next_row(ws) + 2

    _create_withdrawals_section(ws, row, data["withdrawals"])
    row = _get_next_row(ws) + 2

    _create_year_summary(ws, row, data)


def _create_disposals_section(ws, start_row, disposals):
    if not disposals:
        return

    ws.merge_cells(f"A{start_row}:I{start_row}")
    ws.cell(row=start_row, column=1, value="COMPRAS Y VENTAS (Ganancias/Pérdidas Patrimoniales)").font = SECTION_FONT

    row = start_row + 1
    headers = ["Fecha Venta", "Moneda", "Cantidad Vendida", "Valor Venta (EUR)", "Fecha Compra", "Coste Adquisición (EUR)", "Ganancia/Pérdida (EUR)", "Holding", "ID Operación"]
    _set_subheader_row(ws, row, headers)

    for d in disposals:
        row += 1
        tx = d["tx"]
        holding_days = (tx["time"] - d["disposals"][0]["acquisition_date"]).days if d["disposals"] else 0

        values = [
            tx["time"].strftime("%d/%m/%Y"),
            tx["currency"],
            f"{abs(Decimal(str(tx['change']))):.8f}",
            float(d["total_sale_value_eur"]),
            d["disposals"][0]["acquisition_date"].strftime("%d/%m/%Y") if d["disposals"] else "",
            float(d["total_acquisition_cost_eur"]),
            float(d["gain_loss_eur"]),
            f"{holding_days} días",
            str(tx.get("tx_id", "")),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_cell(cell)
            if col in (4, 6, 7):
                cell.number_format = '#,##0.00 "€"'
                if col == 7:
                    cell.font = POSITIVE_FONT if val >= 0 else NEGATIVE_FONT
                    cell.fill = GAINS_FILL if val >= 0 else LOSSES_FILL


def _create_incomes_section(ws, start_row, incomes):
    if not incomes:
        return

    ws.merge_cells(f"A{start_row}:I{start_row}")
    ws.cell(row=start_row, column=1, value="RENDIMIENTOS DE CAPITAL MOBILIARIO (Intereses, Staking, Airdrops)").font = SECTION_FONT

    row = start_row + 1
    headers = ["Fecha", "Tipo", "Moneda", "Cantidad", "Precio EUR", "Valor (EUR)", "Observación"]
    _set_subheader_row(ws, row, headers)

    for inc in incomes:
        row += 1
        values = [
            inc["time"].strftime("%d/%m/%Y"),
            inc["operation"],
            inc["currency"],
            f"{Decimal(str(inc['change'])):.8f}",
            float(inc.get("eur_price") or 0),
            float(inc.get("amount_eur") or 0),
            inc.get("observation", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_cell(cell)
            if col in (5, 6):
                cell.number_format = '#,##0.00 "€"'


def _create_deposits_section(ws, start_row, deposits):
    if not deposits:
        return

    ws.merge_cells(f"A{start_row}:I{start_row}")
    ws.cell(row=start_row, column=1, value="DEPÓSITOS").font = SECTION_FONT

    row = start_row + 1
    headers = ["Fecha", "Moneda", "Cantidad", "Valor (EUR)", "Observación"]
    _set_subheader_row(ws, row, headers)

    for dep in deposits:
        row += 1
        values = [
            dep["time"].strftime("%d/%m/%Y"),
            dep["currency"],
            f"{Decimal(str(dep['change'])):.8f}",
            float(dep.get("amount_eur") or 0),
            dep.get("observation", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_cell(cell)
            if col == 4:
                cell.number_format = '#,##0.00 "€"'


def _create_withdrawals_section(ws, start_row, withdrawals):
    if not withdrawals:
        return

    ws.merge_cells(f"A{start_row}:I{start_row}")
    ws.cell(row=start_row, column=1, value="RETIRADAS").font = SECTION_FONT

    row = start_row + 1
    headers = ["Fecha", "Moneda", "Cantidad", "Valor (EUR)", "Observación"]
    _set_subheader_row(ws, row, headers)

    for w in withdrawals:
        row += 1
        values = [
            w["time"].strftime("%d/%m/%Y"),
            w["currency"],
            f"{abs(Decimal(str(w['change']))):.8f}",
            float(w.get("amount_eur") or 0),
            w.get("observation", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_cell(cell)
            if col == 4:
                cell.number_format = '#,##0.00 "€"'


def _create_year_summary(ws, start_row, data):
    ws.merge_cells(f"A{start_row}:I{start_row}")
    ws.cell(row=start_row, column=1, value="RESUMEN EJERCICIO").font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws.cell(row=start_row, column=1).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    row = start_row + 1
    net = data["net_gain_loss"]
    income = data["income_value_eur"]
    tax = _compute_spanish_tax(net, income)

    summary_items = [
        ("Ganancias patrimoniales:", format_eur(data["total_gains"])),
        ("Pérdidas patrimoniales:", format_eur(data["total_losses"])),
        ("Ganancia/Pérdida neta:", format_eur(net)),
        ("", ""),
        ("Rendimientos capital mobiliario:", format_eur(income)),
        ("", ""),
        ("Impuesto estimado sobre ganancias:", format_eur(tax[0])),
        ("Impuesto estimado sobre rendimientos:", format_eur(tax[1])),
        ("Impuesto total estimado:", format_eur(tax[0] + tax[1])),
    ]

    for label, value in summary_items:
        if label:
            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            ws.cell(row=row, column=2, value=value).font = BOLD_FONT
        row += 1


def _create_inventory_sheet(ws, inventory):
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18

    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="INVENTARIO ACTUAL DE CRIPTOMONEDAS")
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    subtitle = ws.cell(row=2, column=1, value="Saldo pendiente de venta según método FIFO")
    subtitle.font = NORMAL_FONT
    subtitle.alignment = Alignment(horizontal="center")

    row = 4
    _set_header_row(ws, row, ["Moneda", "Cantidad", "Coste Total (EUR)", "Precio Medio (EUR)", "Lote más antiguo", "Total Lotes"])

    grand_total_cost = Decimal("0")
    for currency, lots in sorted(inventory.items()):
        if currency.upper() == "EUR":
            continue
        row += 1
        total_qty = sum(lot["quantity"] for lot in lots)
        total_cost = sum(lot["cost_basis"] for lot in lots)
        avg_price = total_cost / total_qty if total_qty else Decimal("0")
        oldest_date = min(lot["date"] for lot in lots)
        grand_total_cost += total_cost

        values = [currency, f"{total_qty:.8f}", float(total_cost), float(avg_price), oldest_date.strftime("%d/%m/%Y"), len(lots)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_cell(cell)
            if col in (3, 4):
                cell.number_format = '#,##0.00 "€"'
            if col == 3:
                cell.font = BOLD_FONT

    row += 2
    ws.cell(row=row, column=1, value="COSTE TOTAL DEL INVENTARIO:").font = Font(name="Arial", bold=True, size=12, color="2F5496")
    cost_cell = ws.cell(row=row, column=3, value=float(grand_total_cost))
    cost_cell.font = Font(name="Arial", bold=True, size=12, color="2F5496")
    cost_cell.number_format = '#,##0.00 "€"'

    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    note = ws.cell(row=row, column=1, value="NOTA: Este inventario muestra las criptomonedas que aún no han sido vendidas. El coste total es el valor que se utilizará como base fiscal cuando se vendan.")
    note.font = Font(name="Arial", italic=True, size=9, color="808080")


def _get_next_row(ws):
    return ws.max_row


def _compute_spanish_tax(gain_loss, income):
    gain_tax = _apply_savings_brackets(gain_loss) if gain_loss > 0 else Decimal("0")
    income_tax = _apply_savings_brackets(income) if income > 0 else Decimal("0")
    return (gain_tax, income_tax)


def _apply_savings_brackets(amount):
    if amount <= 0:
        return Decimal("0")

    tax = Decimal("0")
    remaining = amount

    brackets = [
        (Decimal("6000"), Decimal("0.19")),
        (Decimal("44000"), Decimal("0.21")),
        (Decimal("150000"), Decimal("0.23")),
        (Decimal("1800000"), Decimal("0.27")),
        (None, Decimal("0.28")),
    ]

    for limit, rate in brackets:
        if remaining <= 0:
            break
        if limit is None:
            tax += remaining * rate
            remaining = Decimal("0")
        else:
            taxable = min(remaining, limit)
            tax += taxable * rate
            remaining -= taxable

    return tax
